import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from collections import Counter

read_file_path = "./lotto.xlsx"
create_file_path = "./example.xlsx"
data = pd.read_excel(read_file_path)

if os.path.exists(create_file_path):
    os.remove(create_file_path)
    
# 연속된 패턴 저장
patterns = []

# 각 행에서 연속된 숫자 패턴 추출
for row in data.values:
    row_list = list(row)  # 행 데이터를 리스트로 변환
    for i in range(len(row_list) - 1):  # 모든 연속된 숫자 쌍 추출
        for j in range(i + 2, len(row_list) + 1):  # 길이가 2 이상인 모든 연속 패턴
            patterns.append(tuple(row_list[i:j]))

# 패턴 빈도 계산
pattern_counts = Counter(patterns)

# 워크북 생성
wb = Workbook()

# 기본 시트 삭제 (기본적으로 생성되는 시트)
del wb['Sheet']

# 첫 번째 시트 ('패턴순위') 생성
ws = wb.create_sheet('패턴순위')

# 헤더 스타일 (배경색 + 가운데 정렬)
header_fill = PatternFill(start_color="3D5A80", end_color="3D5A80", fill_type="solid")  # 노란색 배경
header_alignment = Alignment(horizontal='center', vertical='center')  # 가운데 정렬
header_font = Font(color="FFFFFF")  # 하얀색 글자

# 헤더 추가 (1열: 패턴, 2열: 빈도수)
ws['A1'] = '패턴'
ws['B1'] = '빈도수'

# 헤더에 스타일 적용
ws['A1'].fill = header_fill
ws['B1'].fill = header_fill
ws['A1'].alignment = header_alignment
ws['B1'].alignment = header_alignment
ws['A1'].font = header_font
ws['B1'].font = header_font


# 빈도수 기준으로 정렬
sorted_patterns = pattern_counts.most_common()

# 엑셀에 결과 기록 (1열: 패턴, 2열: 빈도수)
for index, (pattern, count) in enumerate(sorted_patterns, start=2):
    # tuple을 공백으로 구분하여 저장
    pattern_str = ' '.join(map(str, pattern))  # 숫자들을 공백으로 구분하여 문자열로 변환
    ws[f'A{index}'] = pattern_str  # 패턴을 공백으로 구분하여 저장
    ws[f'A{index}'].alignment = header_alignment
    ws[f'B{index}'] = count  # 빈도수
    ws[f'B{index}'].alignment = header_alignment  # 빈도수

# 두 번째 시트 ('숫자순위') 생성
ws2 = wb.create_sheet('숫자순위')

# 헤더 추가 (1열: 번호, 2열: 빈도수)
ws2['A1'] = '번호'
ws2['B1'] = '빈도수'

# 헤더에 스타일 적용
ws2['A1'].fill = header_fill
ws2['B1'].fill = header_fill
ws2['A1'].alignment = header_alignment
ws2['B1'].alignment = header_alignment
ws2['A1'].font = header_font
ws2['B1'].font = header_font

# 데이터를 1차원 배열로 평탄화한 후 빈도 계산
flattened_data = data.values.flatten()
frequency = pd.Series(flattened_data).value_counts()

# 빈도수 기록
for index, (number, count) in enumerate(frequency.items(), start=2):
    ws2[f'A{index}'] = number
    ws2[f'A{index}'].alignment = header_alignment
    ws2[f'B{index}'] = count
    ws2[f'B{index}'].alignment = header_alignment

# 파일 저장
wb.save(create_file_path)